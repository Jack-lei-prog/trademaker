"""聊天 Blueprint — /api/chat, /api/chat/stream, /api/clear, /"""
from flask import Blueprint, render_template, request, jsonify, Response
from security import rate_limit
from services import _safe_str, _needs_tools, run_agent, run_agent_stream
import db
import json

chat_bp = Blueprint("chat", __name__)


@chat_bp.route("/")
def index():
    return render_template("index.html")


@chat_bp.route("/api/chat", methods=["POST"])
@rate_limit(max_requests=30, window=60)
def chat():
    data = request.get_json()
    user_input = _safe_str(data.get("message")).strip()
    user_email = _safe_str(data.get("user_email")).strip().lower()
    session_id = user_email or _safe_str(data.get("session_id")) or "default"

    if not user_input:
        return jsonify({"error": "请输入消息"}), 400

    if user_input.lower() in ("quit", "exit", "q", "退出"):
        return jsonify({"reply": "感谢使用外贸通，再见！👋", "tool_calls": []})

    if user_input.lower() in ("help", "h", "帮助"):
        return jsonify({"reply": get_help_text(), "tool_calls": []})

    if user_input.lower() in ("clear", "清空", "新会话"):
        db.delete_session(user_email, session_id)
        return jsonify({"reply": "会话已清空，我们可以开始新的对话了！", "tool_calls": []})

    result = run_agent(user_input, session_id, use_tools=_needs_tools(user_input), user_email=user_email)
    return jsonify(result)


@chat_bp.route("/api/chat/stream", methods=["POST"])
def chat_stream():
    data = request.get_json()
    user_input = _safe_str(data.get("message")).strip()
    user_email = _safe_str(data.get("user_email")).strip().lower()
    session_id = user_email or _safe_str(data.get("session_id")) or "default"

    if not user_input:
        def eg(): yield "data: {\"type\":\"error\",\"message\":\"请输入消息\"}\n\n"
        return Response(eg(), mimetype="text/event-stream")

    if user_input.lower() in ("quit", "exit", "q", "退出"):
        def qg():
            yield f"data: {json.dumps({'type': 'text', 'content': '感谢使用外贸通，再见！👋'}, ensure_ascii=False)}\n\n"
            yield "data: {\"type\":\"done\"}\n\n"
        return Response(qg(), mimetype="text/event-stream")

    if user_input.lower() in ("help", "h", "帮助"):
        help_text = get_help_text()
        def hg():
            yield f"data: {json.dumps({'type': 'text', 'content': help_text}, ensure_ascii=False)}\n\n"
            yield "data: {\"type\":\"done\"}\n\n"
        return Response(hg(), mimetype="text/event-stream")

    if user_input.lower() in ("clear", "清空", "新会话"):
        db.delete_session(user_email, session_id)
        def cg():
            yield f"data: {json.dumps({'type': 'text', 'content': '会话已清空！'}, ensure_ascii=False)}\n\n"
            yield "data: {\"type\":\"done\"}\n\n"
        return Response(cg(), mimetype="text/event-stream")

    def generate():
        for event in run_agent_stream(user_input, session_id, use_tools=_needs_tools(user_input), user_email=user_email):
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
        yield "data: {\"type\":\"done\"}\n\n"
    return Response(generate(), mimetype="text/event-stream")


@chat_bp.route("/api/clear", methods=["POST"])
def clear_chat():
    data = request.get_json() or {}
    session_id = _safe_str(data.get("session_id")) or "default"
    user_email = _safe_str(data.get("user_email"))
    db.delete_session(user_email=user_email, session_id=session_id)
    return jsonify({"success": True})


@chat_bp.route("/api/health", methods=["GET"])
def health():
    """健康检查：数据库状态、API 连通性、运行时间"""
    import time as _t
    import os as _os
    status = {"status": "ok", "service": "TradeMaster", "version": "2.1.0"}
    checks = {}

    # 数据库检查
    try:
        db.get_or_create_session("health_check", "health")
        checks["database"] = "ok"
    except Exception as e:
        checks["database"] = f"error: {str(e)[:100]}"
        status["status"] = "degraded"

    # API 连通性检查 + 自动修复
    from services import get_provider_status, LLM_PROVIDERS, reset_all_providers
    statuses = {p["name"]: get_provider_status().get(p["name"], {}).get("status", "unknown")
                for p in LLM_PROVIDERS}
    # 如果所有provider都在错误状态，自动重置（API可能已恢复）
    all_error = all("error" in str(v).lower() for v in statuses.values())
    if all_error and LLM_PROVIDERS:
        reset_all_providers()
        statuses = {p["name"]: "reset" for p in LLM_PROVIDERS}
    checks["llm_api"] = {"providers": len(LLM_PROVIDERS), "detail": statuses}

    # 数据源状态检查
    from data_sources import DATA_SOURCE_STATUS
    checks["data_sources"] = dict(DATA_SOURCE_STATUS)

    # 磁盘检查
    try:
        stat = _os.statvfs(".") if hasattr(_os, "statvfs") else None
        if stat:
            free_mb = stat.f_frsize * stat.f_bavail / (1024 * 1024)
            checks["disk_free_mb"] = round(free_mb, 1)
    except Exception:
        pass

    status["checks"] = checks
    status["timestamp"] = _t.strftime("%Y-%m-%d %H:%M:%S", _t.localtime())
    return jsonify(status)


@chat_bp.route("/api/demo/help", methods=["GET"])
def demo_help():
    """返回演示指南"""
    from knowledge.demo import DEMO_HELP, DEMO_SCENARIOS
    return jsonify({
        "success": True,
        "demo_account_email": "demo@trademaster.com",
        "demo_account_hint": "密码为 'demo2024'（仅演示环境），请访问登录页面使用该账号体验",
        "scenarios": DEMO_SCENARIOS,
        "help_text": DEMO_HELP,
    })


@chat_bp.route("/api/docs", methods=["GET"])
def api_docs():
    """API 文档页面"""
    docs = {
        "service": "TradeMaster 外贸通",
        "version": "2.1.0",
        "architecture": "单Agent + 12 Skill协同 (Function Calling) + RAG知识检索",
        "endpoints": [
            {"method": "POST", "path": "/api/register", "desc": "用户注册", "auth": False},
            {"method": "POST", "path": "/api/login", "desc": "用户登录", "auth": False},
            {"method": "POST", "path": "/api/chat", "desc": "AI对话 (同步)", "auth": False, "stream": False},
            {"method": "POST", "path": "/api/chat/stream", "desc": "AI对话 (SSE流式)", "auth": False, "stream": True},
            {"method": "POST", "path": "/api/clear", "desc": "清空会话", "auth": False},
            {"method": "GET", "path": "/api/health", "desc": "健康检查+API状态", "auth": False},
            {"method": "GET", "path": "/api/docs", "desc": "本API文档", "auth": False},
            {"method": "GET", "path": "/api/demo/help", "desc": "演示指南", "auth": False},
            {"method": "POST", "path": "/api/dashboard", "desc": "仪表盘数据", "auth": True},
            {"method": "POST", "path": "/api/send_email", "desc": "生成邮件发送界面", "auth": False},
            {"method": "POST", "path": "/api/email/smtp_send", "desc": "SMTP发送邮件", "auth": False},
            {"method": "POST", "path": "/api/email/stats", "desc": "邮件统计", "auth": True},
            {"method": "POST", "path": "/api/emails/sent", "desc": "已发送邮件列表", "auth": True},
            {"method": "POST", "path": "/api/emails/pending", "desc": "待跟进邮件", "auth": True},
            {"method": "POST", "path": "/api/contacts/add", "desc": "添加待联系客户", "auth": True},
            {"method": "POST", "path": "/api/contacts/list", "desc": "客户列表", "auth": True},
            {"method": "POST", "path": "/api/contacts/update", "desc": "更新客户状态", "auth": True},
            {"method": "POST", "path": "/api/contacts/stats", "desc": "客户统计", "auth": True},
            {"method": "POST", "path": "/api/inquiry/process", "desc": "询盘处理", "auth": False},
            {"method": "POST", "path": "/api/evaluate", "desc": "评价回答质量", "auth": False},
            {"method": "POST", "path": "/api/evaluate/kimi", "desc": "Kimi深度评价", "auth": False},
        ],
        "skills": [
            "buyer_search — 买家搜索 (Wikidata/OpenCorp/LLM)",
            "email_draft — 开发信撰写 (B2B模板+企业拦截)",
            "trade_intelligence — 展会情报 (50+展会+认证+术语)",
            "inquiry_processing — 询盘处理 (5步闭环)",
            "email_tracking — 邮件追踪 (像素+退信检测)",
            "contact_management — 客户管理 (7状态Pipeline)",
        ],
        "knowledge_base": {
            "tradeshows": "50+全球展会数据库，按产品匹配",
            "certifications": "CE/FCC/RoHS/BQB等出口认证清单",
            "trade_terms": "18个外贸术语 (FOB/CIF/MOQ/OEM等)",
            "rag_search": "TF-IDF语义检索，/api/chat自动调用"
        }
    }
    return jsonify(docs)


@chat_bp.route("/api/upload/manual", methods=["POST"])
@rate_limit(max_requests=10, window=300)
def upload_manual():
    """上传产品手册（PDF/DOCX/TXT），提取文本存入会话"""
    import os, tempfile
    file = request.files.get("file")
    user_email = request.form.get("user_email", "").strip().lower()
    session_id = request.form.get("session_id", user_email or "default")

    if not file:
        return jsonify({"success": False, "error": "请选择文件"}), 400

    # 文件大小限制 10MB
    file_content = file.read()
    if len(file_content) > 10 * 1024 * 1024:
        return jsonify({"success": False, "error": "文件大小不能超过 10MB"}), 400
    from io import BytesIO
    file = BytesIO(file_content)  # 供下游 PdfReader 使用

    # 保存临时文件
    ext = os.path.splitext(file.filename or "manual.pdf")[1].lower()
    if ext not in (".pdf", ".txt", ".docx"):
        return jsonify({"success": False, "error": f"不支持的格式: {ext}，支持 PDF/TXT/DOCX"}), 400

    try:
        text = ""
        if ext == ".pdf":
            from PyPDF2 import PdfReader
            import io
            reader = PdfReader(file)
            for page in reader.pages:
                t = page.extract_text()
                if t: text += t + "\n"
        elif ext == ".txt":
            text = file.read().decode("utf-8", errors="replace")
        elif ext == ".docx":
            return jsonify({"success": False, "error": "DOCX暂不支持，请先转成PDF或TXT"}), 400

        if not text.strip():
            return jsonify({"success": False, "error": "无法从文件中提取文字"}), 400

        # 截取前 8000 字符存入会话
        text = text.strip()[:8000]
        word_count = len(text)

        # 存入会话 metadata
        metadata = db.get_session_metadata(user_email, session_id)
        metadata["product_manual"] = text
        metadata["manual_filename"] = file.filename
        metadata["manual_uploaded_at"] = __import__('db')._now()
        db.update_session_metadata(user_email, session_id, metadata)

        return jsonify({
            "success": True,
            "message": f"已解析 {file.filename}，共 {word_count} 字符",
            "filename": file.filename,
            "preview": text[:300] + ("..." if len(text) > 300 else ""),
        })
    except Exception as e:
        return jsonify({"success": False, "error": f"解析失败: {str(e)}"}), 500


@chat_bp.route("/api/upload/excel", methods=["POST"])
@rate_limit(max_requests=10, window=300)
def upload_excel():
    """上传厂家Excel表格，解析出厂家列表"""
    import io, os as _os
    file = request.files.get("file")
    user_email = request.form.get("user_email", "").strip().lower()

    if not file:
        return jsonify({"success": False, "error": "请选择文件"}), 400

    # 文件大小限制 10MB
    file_content = file.read()
    if len(file_content) > 10 * 1024 * 1024:
        return jsonify({"success": False, "error": "文件大小不能超过 10MB"}), 400

    ext = _os.path.splitext(file.filename or "list.xlsx")[1].lower()
    if ext not in (".xlsx", ".xls"):
        return jsonify({"success": False, "error": f"不支持{wxt}，请上传 .xlsx 格式"}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active

        # 读取表头（第一行）
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").strip().lower())

        # 映射常见列名到标准字段
        col_map = {}
        name_keys = ["company_name", "company", "name", "公司名", "公司名称", "厂家", "客户名称", "客户"]
        email_keys = ["email", "mail", "e-mail", "邮箱", "邮件", "电子邮件"]
        contact_keys = ["contact", "person", "contact_person", "name", "联系人", "姓名", "负责人"]
        phone_keys = ["phone", "tel", "telephone", "mobile", "电话", "手机", "联系电话"]
        country_keys = ["country", "nation", "国家", "地区"]
        product_keys = ["product", "interest", "产品", "采购产品", "主营", "需求"]
        website_keys = ["website", "web", "url", "site", "网站", "官网"]

        for i, h in enumerate(headers):
            h_lower = h.lower().replace(" ", "").replace("_", "")
            if any(k in h_lower for k in name_keys): col_map["company_name"] = i
            elif any(k in h_lower for k in email_keys): col_map["email"] = i
            elif any(k in h_lower for k in contact_keys): col_map["contact_person"] = i
            elif any(k in h_lower for k in phone_keys): col_map["phone"] = i
            elif any(k in h_lower for k in country_keys): col_map["country"] = i
            elif any(k in h_lower for k in product_keys): col_map["product_interest"] = i
            elif any(k in h_lower for k in website_keys): col_map["website"] = i

        # 如果没识别到 email 列，尝试模糊匹配
        if "email" not in col_map:
            for i, h in enumerate(headers):
                if "@" in str(ws.cell(row=2, column=i+1).value or ""):
                    col_map["email"] = i
                    break

        # 读取数据行
        companies = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            entry = {"company_name": "", "email": "", "contact_person": "",
                     "phone": "", "country": "", "product_interest": "", "website": ""}
            for field, idx in col_map.items():
                val = str(row[idx] or "").strip()
                entry[field] = val
            # 至少要有公司名或邮箱
            if entry["company_name"] or entry["email"]:
                companies.append(entry)

        if not companies:
            return jsonify({"success": False, "error": "未找到有效数据，请检查Excel格式"}), 400

        # 截取前100条
        total = len(companies)
        companies = companies[:100]

        # 存入会话
        import db as _db
        metadata = _db.get_session_metadata(user_email or "", user_email or "default")
        metadata["excel_companies"] = companies
        _db.update_session_metadata(user_email or "", user_email or "default", metadata)

        return jsonify({
            "success": True,
            "total": total,
            "loaded": len(companies),
            "columns": list(col_map.keys()),
            "companies": companies,
            "message": f"已解析 {len(companies)} 条厂家信息（共{total}条，最多加载100条）"
        })
    except Exception as e:
        return jsonify({"success": False, "error": f"解析失败: {str(e)}"}), 500


def get_help_text():
    return """## 📖 使用帮助

**💬 AI对话**
1. 🔍 买家搜索 — "搜索德国蓝牙耳机进口商"
2. ✉️ 开发信撰写 — "给XXX公司写一封开发信"
3. 📩 询盘处理 — 直接粘贴客户询盘邮件
4. 💱 汇率查询 — "美元兑人民币汇率"
5. 📝 商品描述 — "为蓝牙耳机生成商品描述"
6. 📈 销售分析 — "分析今天的销售数据"
7. 🎯 广告语 — "为防晒衣生成广告语"

**📎 文件上传（输入框下方按钮）**
8. 📄 产品手册 — 上传PDF/TXT，开发信和广告语自动引用手册内容
9. 📋 厂家列表 — 上传Excel批量导入客户，逐家发送开发信

**💡 提示：上传产品手册后，Agent写开发信和广告语时会自动引用手册中的产品参数"""
